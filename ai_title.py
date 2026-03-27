import sys
import os
import base64
import requests
import datetime
from pathlib import Path
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton,
    QFileDialog, QTextEdit, QLabel
)
from PyQt6.QtGui import QIcon
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 如果你是为了打包后任务栏能正确显示图标，需要加这一行
try:
    from ctypes import windll
    windll.shell32.SetCurrentProcessExplicitAppUserModelID("my_ai_tool_v1")
except:
    pass

API_KEY = "sk-46X32UEI0hNcKczjbwivYhrlvJgwQjOwCXQ7jZxut7oscoSo"
API_URL = "https://api.moonshot.cn/v1/chat/completions"
MODEL = "moonshot-v1-8k-vision-preview"

PROMPT = """
请识别这张照片的内容（人物关系、场景、情感氛围等），为相框商品生成一个拼多多爆款标题，要求：
1. 描述照片里的温馨场景（如亲子时光、情侣瞬间、家庭团聚等）
2. 突出相框能承载回忆、装点生活的卖点
3. 暗示适用场景（如客厅、卧室、办公桌、礼物赠送）
4. 30字以内，输出中文，不要解释
"""

def image_to_base64(image_path):
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")

def call_kimi_api(base64_image):
    headers = {
        "Authorization": f"Bearer {API_KEY.strip()}",
        "Content-Type": "application/json"
    }
    data = {
        "model": MODEL,
        "messages": [
            {"role": "system", "content": "你是一个电商运营专家"},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": PROMPT},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}}
                ]
            }
        ]
    }
    try:
        response = requests.post(API_URL, headers=headers, json=data, timeout=60)
        result = response.json()
        if "choices" in result:
            return result["choices"][0]["message"]["content"]
        return f"接口错误: {result.get('error', {}).get('message', '未知错误')}"
    except Exception as e:
        return f"请求失败: {e}"

class App(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("AI 电商标题批量生成工具")
        self.resize(800, 600)

        # 加载图标文件，请确保 logo.ico 存在
        self.set_app_icon()

        self.init_ui()

    def set_app_icon(self):
        icon_path = os.path.join(os.path.dirname(__file__), "logo.ico")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

    def init_ui(self):
        layout = QVBoxLayout()

        self.label = QLabel("操作指引：选择文件夹后自动处理，结果保存至同目录 Excel")
        self.label.setStyleSheet("color: #333; font-size: 15px; font-weight: bold; padding: 5px;")
        layout.addWidget(self.label)

        self.btn_run = QPushButton("🚀 选择文件夹并批量生成标题")
        self.btn_run.setFixedHeight(60)
        self.btn_run.setStyleSheet("""
            QPushButton {
                background-color: #0078D7;
                color: white;
                font-size: 18px;
                font-weight: bold;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #106EBE;
            }
        """)
        self.btn_run.clicked.connect(self.process_all)
        layout.addWidget(self.btn_run)

        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setStyleSheet("""
            QTextEdit {
                background-color: #1E1E1E;
                color: #00FF00;
                font-family: 'Consolas', 'Monaco', 'Courier New';
                font-size: 14px;
                border: 2px solid #333;
                padding: 10px;
            }
        """)
        layout.addWidget(self.log)
        self.setLayout(layout)

    def log_msg(self, msg):
        self.log.append(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}")
        self.log.verticalScrollBar().setValue(self.log.verticalScrollBar().maximum())
        QApplication.processEvents()

    def apply_excel_style(self, ws):
        header_font = Font(name='微软雅黑', bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_side = Side(style='thin', color="000000")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="left")
                cell.border = border

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 80

    def process_all(self):
        image_dir = QFileDialog.getExistingDirectory(self, "选择图片文件夹")
        if not image_dir:
            return

        folder_path = Path(image_dir)
        files = [f for f in folder_path.glob("*.*") if f.suffix.lower() in [".jpg", ".jpeg", ".png"]]

        if not files:
            self.log_msg("⚠️ 错误：未发现图片文件")
            return

        self.log.clear()
        self.log_msg(f"📂 目录: {folder_path}")

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        save_path = folder_path / f"商品标题导出_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "AI标题生成"
        ws.append(["原始图片名", "爆款标题"])

        for i, f in enumerate(files, 1):
            self.log_msg(f"处理中({i}/{len(files)}): {f.name}")
            content = call_kimi_api(image_to_base64(f))
            ws.append([f.name, content])

        self.apply_excel_style(ws)

        try:
            wb.save(save_path)
            self.log_msg(f"✅ 完成！文件已保存。")
            os.startfile(save_path)
        except Exception as e:
            self.log_msg(f"❌ 保存异常: {e}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = App()
    window.show()
    sys.exit(app.exec())
